"""
Batch processor for Raw_Work.xlsx files.

Reads order rows, calculates box costs using BoxCostCalculator with
date-based material costs, applies margin, and writes Rate + Total back.

Supports three order types via the Order Type column:
  - All: full cost calculation (paper + corrugation + labour)
  - Corrugation: corrugation processing fee only (no paper cost)
  - Labour: labour costs only (punching, pasting, pins, hand pasting)

Usage (CLI):  python -m calculator.batch_processor Raw_Work.xlsx
Usage (API):  process_workbook(filepath) -> list of result dicts
"""

import math
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

from .box_cost_calculator import (
    BoxCostCalculator,
    BoxDimensions,
    PaperQuality,
    ProductionDetails,
    ManufacturingOptions,
)
from .material_costs import get_costs_for_date

# Column indices (1-based) matching Raw_Work.xlsx layout
COL_NAME           = 1
COL_GROUP          = 2
COL_DATE           = 3
COL_LENGTH         = 4
COL_WIDTH          = 5
COL_BOTTOM_WEIGHT  = 6
COL_BOTTOM_QUALITY = 7
COL_FLUTE_WEIGHT   = 8
COL_FLUTE_QUALITY  = 9
COL_TOP_WEIGHT     = 10
COL_TOP_QUALITY    = 11
COL_PLY            = 12
COL_ORDER_TYPE     = 13
COL_UPS            = 14
COL_NUM_BOXES      = 15
COL_PUNCHING       = 16
COL_PINS           = 17
COL_ITEM_NAME      = 18
COL_RATE           = 19
COL_TOTAL          = 20

# Valid Order Type values
NATURE_ALL         = 'ALL'
NATURE_CORRUGATION = 'CORRUGATION'
NATURE_LABOUR      = 'LABOUR'

# Margin by customer group
MARGIN_MAP = {
    'A': 0.05,
    'B': 0.10,
    'C': 0.15,
    'D': 0.20,
}

# Map Excel quality strings to PaperQuality enum names
QUALITY_MAP = {
    'kraft':         'KRAFT',
    'duplex':        'DUPLEX',
    'golden':        'GOLDEN',
    'preprinted':    'PREPRINTED',
    'golden180':     'GOLDEN180',
    'itc':           'ITC',
}

# Headers for the Raw_Work template (input columns only;
# Rate/Total are written to cols 19-20 by the processor).
HEADERS = [
    'Customer Name', 'Group', 'Date', 'Length', 'Width',
    'Bottom', 'Bottom Quality', 'Flute', 'Flute Quality',
    'Top', 'Top Quality', 'Ply', 'Order Type', 'Ups',
    'Number of Boxes', 'Punching', 'Pins', 'Item Name',
]

# Data validation rule range — covers a generous number of input rows
TEMPLATE_VALIDATION_LAST_ROW = 1000


def _lookup_quality(quality_str):
    """Map an Excel quality string to a PaperQuality enum member."""
    key = str(quality_str).strip().lower()
    if key not in QUALITY_MAP:
        raise ValueError(
            f"Unknown paper quality: '{quality_str}'. "
            f"Valid values: {list(QUALITY_MAP.keys())}"
        )
    return PaperQuality[QUALITY_MAP[key]]


def _date_to_str(date_val):
    """Convert an Excel date cell value to 'YYYY-MM' string, or None."""
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val.strftime('%Y-%m')
    # Try parsing common string formats
    s = str(date_val).strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y-%m'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m')
        except ValueError:
            continue
    return None


def _add_template_validations(ws):
    """Attach dropdowns and data-type validations to every input column.

    Applies to data rows 2..TEMPLATE_VALIDATION_LAST_ROW. Each rule shows a
    Stop-style error so invalid entries are rejected by Excel.
    """
    last = TEMPLATE_VALIDATION_LAST_ROW

    def _add(rule, col_idx):
        rule.showErrorMessage = True
        rule.errorStyle = 'stop'
        rule.showInputMessage = True
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        rule.add(f'{col_letter}2:{col_letter}{last}')
        ws.add_data_validation(rule)

    # Group: A/B/C/D
    _add(DataValidation(
        type='list', formula1='"A,B,C,D"',
        error='Group must be A, B, C, or D.',
        promptTitle='Group', prompt='Customer group (A=5%, B=10%, C=15%, D=20% margin).',
    ), COL_GROUP)

    # Date: must be a valid date
    _add(DataValidation(
        type='date', operator='greaterThanOrEqual', formula1='1900-01-01',
        error='Enter a valid date.',
        promptTitle='Date', prompt='Order date (used for material cost lookup).',
    ), COL_DATE)

    # Length: 18-32 inches
    _add(DataValidation(
        type='decimal', operator='between', formula1=18, formula2=32,
        error='Length must be between 18 and 32 inches.',
        promptTitle='Length', prompt='Sheet length in inches (18-32).',
    ), COL_LENGTH)

    # Width: positive decimal
    _add(DataValidation(
        type='decimal', operator='greaterThan', formula1=0,
        error='Width must be a positive number.',
        promptTitle='Width', prompt='Sheet width in inches (any positive value).',
    ), COL_WIDTH)

    # Bottom weight: whole number > 0
    _add(DataValidation(
        type='whole', operator='greaterThan', formula1=0,
        error='Bottom weight (GSM) must be a whole number greater than 0.',
        promptTitle='Bottom (GSM)', prompt='Bottom paper weight in GSM (whole number > 0).',
    ), COL_BOTTOM_WEIGHT)

    # Bottom quality: Kraft/Duplex/Golden (no Preprinted, no ITC)
    _add(DataValidation(
        type='list', formula1='"Kraft,Duplex,Golden"',
        error='Bottom Quality must be Kraft, Duplex, or Golden.',
        promptTitle='Bottom Quality', prompt='Bottom layer paper quality.',
    ), COL_BOTTOM_QUALITY)

    # Flute weight: whole number > 0
    _add(DataValidation(
        type='whole', operator='greaterThan', formula1=0,
        error='Flute weight (GSM) must be a whole number greater than 0.',
        promptTitle='Flute (GSM)', prompt='Flute paper weight in GSM (whole number > 0).',
    ), COL_FLUTE_WEIGHT)

    # Flute quality: Kraft/Duplex/Golden
    _add(DataValidation(
        type='list', formula1='"Kraft,Duplex,Golden"',
        error='Flute Quality must be Kraft, Duplex, or Golden.',
        promptTitle='Flute Quality', prompt='Flute layer paper quality.',
    ), COL_FLUTE_QUALITY)

    # Top weight: whole number >= 0 (0 valid for Preprinted)
    _add(DataValidation(
        type='whole', operator='greaterThanOrEqual', formula1=0,
        error='Top weight (GSM) must be a whole number (0 or greater).',
        promptTitle='Top (GSM)', prompt='Top paper weight in GSM (use 0 for Preprinted).',
    ), COL_TOP_WEIGHT)

    # Top quality: all five
    _add(DataValidation(
        type='list', formula1='"Kraft,Duplex,Golden,ITC,Preprinted"',
        error='Top Quality must be Kraft, Duplex, Golden, ITC, or Preprinted.',
        promptTitle='Top Quality', prompt='Top layer paper quality.',
    ), COL_TOP_QUALITY)

    # Ply: odd, max 9
    _add(DataValidation(
        type='list', formula1='"3,5,7,9"',
        error='Ply must be 3, 5, 7, or 9.',
        promptTitle='Ply', prompt='Ply count (odd, max 9).',
    ), COL_PLY)

    # Order Type: All/Corrugation/Labour
    _add(DataValidation(
        type='list', formula1='"All,Corrugation,Labour"',
        error='Order Type must be All, Corrugation, or Labour.',
        promptTitle='Order Type', prompt='All = full cost; Corrugation = no paper; Labour = labour only.',
    ), COL_ORDER_TYPE)

    # Ups: decimal > 0
    _add(DataValidation(
        type='decimal', operator='greaterThan', formula1=0,
        error='Ups must be a positive number.',
        promptTitle='Ups', prompt='Boxes per sheet (positive, decimals allowed).',
    ), COL_UPS)

    # Number of Boxes: whole number >= 100
    _add(DataValidation(
        type='whole', operator='greaterThanOrEqual', formula1=100,
        error='Number of Boxes must be a whole number of at least 100.',
        promptTitle='Number of Boxes', prompt='Total quantity (minimum 100).',
    ), COL_NUM_BOXES)

    # Punching: Y/N
    _add(DataValidation(
        type='list', formula1='"Y,N"',
        error='Punching must be Y or N.',
        promptTitle='Punching', prompt='Y if the box requires punching, otherwise N.',
    ), COL_PUNCHING)

    # Pins: whole number 0-20
    _add(DataValidation(
        type='whole', operator='between', formula1=0, formula2=20,
        error='Pins must be a whole number between 0 and 20.',
        promptTitle='Pins', prompt='Pins per box (0-20). Leave blank for 0.',
    ), COL_PINS)


def generate_template(filepath):
    """Generate an empty Raw_Work template with headers, samples, and validation."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Write headers
    for col, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col).value = header

    # Sample rows — one per Order Type
    samples = [
        # All: full calculation
        ['Acme Corp', 'B', '2026-03-01', 20, 27, 180, 'Duplex', 100, 'Kraft',
         0, 'Preprinted', 3, 'All', 2, 1000, 'Y', 0, 'Widget Box'],
        # Corrugation: no paper cost, corrugation included
        ['Beta Ltd', 'C', '2026-03-01', 24, 16, 100, 'Kraft', 100, 'Kraft',
         0, 'Preprinted', 3, 'Corrugation', 2, 500, 'Y', 0, ''],
        # Labour: only punching, pasting, pins, hand pasting
        ['Gamma Inc', 'B', '2026-03-01', 19, 16, 150, 'Golden', 150, 'Golden',
         150, 'Golden', 3, 'Labour', 2, 2000, 'Y', 3, 'Punch #305'],
    ]

    for row_idx, sample in enumerate(samples, 2):
        for col_idx, val in enumerate(sample, 1):
            ws.cell(row=row_idx, column=col_idx).value = val

    _add_template_validations(ws)

    wb.save(filepath)
    return filepath


def process_workbook(filepath):
    """
    Process a Raw_Work.xlsx file: calculate costs and write Rate + Total.

    Uses the order date in each row to look up the correct monthly material
    costs. If no date is present, uses the latest available costs.

    Returns a list of dicts with per-row results for API consumption.
    """
    calculator = BoxCostCalculator()
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    results = []

    for row_num in range(2, ws.max_row + 1):
        name = ws.cell(row=row_num, column=COL_NAME).value
        if name is None:
            break

        # Read inputs
        group          = str(ws.cell(row=row_num, column=COL_GROUP).value).strip().upper()
        date_val       = ws.cell(row=row_num, column=COL_DATE).value
        length         = float(ws.cell(row=row_num, column=COL_LENGTH).value)
        width          = float(ws.cell(row=row_num, column=COL_WIDTH).value)
        bottom_weight  = int(ws.cell(row=row_num, column=COL_BOTTOM_WEIGHT).value)
        bottom_quality = _lookup_quality(ws.cell(row=row_num, column=COL_BOTTOM_QUALITY).value)
        flute_weight   = int(ws.cell(row=row_num, column=COL_FLUTE_WEIGHT).value)
        flute_quality  = _lookup_quality(ws.cell(row=row_num, column=COL_FLUTE_QUALITY).value)
        top_weight     = int(ws.cell(row=row_num, column=COL_TOP_WEIGHT).value)
        top_quality    = _lookup_quality(ws.cell(row=row_num, column=COL_TOP_QUALITY).value)
        ply_num        = int(ws.cell(row=row_num, column=COL_PLY).value)
        order_type_val     = str(ws.cell(row=row_num, column=COL_ORDER_TYPE).value).strip().upper()
        box_per_sheet  = float(ws.cell(row=row_num, column=COL_UPS).value)
        number_of_boxes = int(ws.cell(row=row_num, column=COL_NUM_BOXES).value)
        punching_val   = str(ws.cell(row=row_num, column=COL_PUNCHING).value).strip().upper()
        pins_per_box   = int(ws.cell(row=row_num, column=COL_PINS).value or 0)
        item_name      = ws.cell(row=row_num, column=COL_ITEM_NAME).value

        # Derived flags
        is_punching = (punching_val == 'Y')
        only_corrugation = (order_type_val == NATURE_CORRUGATION)
        labour_only = (order_type_val == NATURE_LABOUR)
        # Hand pasting: apply when pins are 0 (for labour/corrugation orders)
        is_hand_pasted = (pins_per_box == 0 and labour_only)

        # Look up material costs for the order date
        date_str = _date_to_str(date_val)
        cost_overrides = get_costs_for_date(date_str)

        # Resolve GOLDEN + 180gsm → GOLDEN180 for pricing
        qualities = [bottom_quality, flute_quality, top_quality]
        weights = [bottom_weight, flute_weight, top_weight]
        for i in range(len(qualities)):
            if qualities[i].name == 'GOLDEN' and weights[i] == 180:
                qualities[i] = PaperQuality.GOLDEN180

        # Build calculator inputs
        production_details = ProductionDetails(
            ply_num=ply_num,
            box_per_sheet=box_per_sheet,
            pins_per_box=pins_per_box,
            number_of_boxes=number_of_boxes,
            paper_weight=weights,
            paper_quality=qualities,
        )

        manufacturing_options = ManufacturingOptions(
            is_punching=is_punching,
            is_scoring=False,
            is_laminated=False,
            is_printed=False,
            is_hand_pasted=is_hand_pasted,
            is_nf=False,
            only_corrugation=only_corrugation,
            labour_only=labour_only,
        )

        # Calculate cost
        result = calculator.calculate_total_cost(
            sheet_length=length,
            sheet_width=width,
            production_details=production_details,
            manufacturing_options=manufacturing_options,
            cost_overrides=cost_overrides,
        )

        cost_per_box = result.manufacturing_cost_per_box

        # Apply margin and round to nearest 0.25
        margin = MARGIN_MAP.get(group, 0.10)
        rate = cost_per_box * (1 + margin)
        rate_value = math.ceil(rate / 0.25) * 0.25
        total_value = round(number_of_boxes * rate_value, 2)

        # Write back to Excel
        ws.cell(row=row_num, column=COL_RATE).value = rate_value
        ws.cell(row=row_num, column=COL_TOTAL).value = total_value

        results.append({
            'row': row_num,
            'name': str(name),
            'group': group,
            'date': date_str or 'latest',
            'order_type': order_type_val,
            'item_name': str(item_name) if item_name else None,
            'cost_per_box': round(cost_per_box, 4),
            'margin': f"{int(margin * 100)}%",
            'rate': rate_value,
            'total': total_value,
        })

    wb.save(filepath)
    return results
