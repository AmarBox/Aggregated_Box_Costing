"""
Transfer processed rows from Raw_Work.xlsx into Estimates.xlsx.

Each row is placed into a sheet named after the customer. Transferred rows
are removed from Raw_Work.xlsx.

Usage (CLI):  python -m calculator.estimate_transfer Raw_Work.xlsx Estimates.xlsx
Usage (API):  transfer(raw_work_path, estimates_path) -> summary dict
"""

from datetime import date, datetime

import openpyxl
from openpyxl.utils import get_column_letter

# Raw_Work column indices (1-based)
RW_NAME       = 1
RW_GROUP      = 2
RW_DATE       = 3
RW_LENGTH     = 4
RW_WIDTH      = 5
RW_BOTTOM_W   = 6
RW_BOTTOM_Q   = 7
RW_FLUTE_W    = 8
RW_FLUTE_Q    = 9
RW_TOP_W      = 10
RW_TOP_Q      = 11
RW_PLY        = 12
RW_ORDER_TYPE = 13
RW_UPS        = 14
RW_NUM_BOXES  = 15
RW_PUNCHING   = 16
RW_PINS       = 17
RW_ITEM_NAME  = 18
RW_RATE       = 19
RW_TOTAL      = 20

# Estimates column indices (1-based)
EST_SR_NO     = 1
EST_DATE      = 2
EST_NAME      = 3
EST_SIZE      = 4
EST_PROGRAM   = 5
EST_PLY       = 6
EST_RATE      = 7
EST_NUM_BOXES = 8
EST_TOTAL     = 9

HEADERS = ['Sr. No.', 'Date', 'Name', 'Size', 'Program', 'Ply', 'Rate', 'Number of Boxes', 'Total']


def _format_num(n):
    """Show as int if whole number, else as float."""
    if float(n) == int(float(n)):
        return str(int(float(n)))
    return str(float(n))


def _build_program(bottom_w, bottom_q, flute_w, flute_q,
                   top_w, top_q, ups, punching, pins, order_type=None):
    """Build a human-readable program description string."""
    prog = f"{bottom_w}({bottom_q}) + {flute_w}({flute_q}) + {top_w}({top_q})"
    prog += f" ; Ups = {_format_num(ups)}"
    if order_type and str(order_type).strip().upper() != 'ALL':
        prog += f" ; {order_type}"
    if str(punching).strip().upper() == 'Y':
        prog += " ; Punching"
    if pins and int(pins) > 0:
        prog += f" ; Pins = {int(pins)}"
    return prog


def _build_size(length, width):
    """Build a size string like '20x27'."""
    return f"{_format_num(length)}x{_format_num(width)}"


def _normalize_date(date_val):
    """Convert a date cell value to a date object for sorting."""
    if isinstance(date_val, datetime):
        return date_val.date()
    if isinstance(date_val, date):
        return date_val
    if isinstance(date_val, str):
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y'):
            try:
                return datetime.strptime(date_val, fmt).date()
            except ValueError:
                pass
    return date.min


def _get_or_create_sheet(wb, name):
    """Return (worksheet, is_new). Creates sheet with headers if it doesn't exist."""
    if name in wb.sheetnames:
        return wb[name], False
    ws = wb.create_sheet(title=name)
    for col, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col).value = header
    return ws, True


def _read_existing_rows(ws):
    """Read all data rows from a sheet into a list of tuples (date, row_data_dict).

    Skips the header row and any Grand Total row.
    """
    rows = []
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row, column=EST_DATE).value
        # Skip Grand Total rows
        total_check = ws.cell(row=row, column=EST_NUM_BOXES).value
        if isinstance(total_check, str) and 'Grand Total' in total_check:
            continue
        # Skip empty rows
        if date_val is None and ws.cell(row=row, column=EST_PROGRAM).value is None:
            continue
        row_data = {
            'date': date_val,
            'name': ws.cell(row=row, column=EST_NAME).value,
            'size': ws.cell(row=row, column=EST_SIZE).value,
            'program': ws.cell(row=row, column=EST_PROGRAM).value,
            'ply': ws.cell(row=row, column=EST_PLY).value,
            'rate': ws.cell(row=row, column=EST_RATE).value,
            'num_boxes': ws.cell(row=row, column=EST_NUM_BOXES).value,
            'total': ws.cell(row=row, column=EST_TOTAL).value,
        }
        rows.append((_normalize_date(date_val), row_data))
    return rows


def _merge_sorted(existing, new_rows):
    """Merge two sorted lists of (date, row_data) into one sorted list."""
    merged = []
    i = j = 0
    while i < len(existing) and j < len(new_rows):
        if existing[i][0] <= new_rows[j][0]:
            merged.append(existing[i])
            i += 1
        else:
            merged.append(new_rows[j])
            j += 1
    merged.extend(existing[i:])
    merged.extend(new_rows[j:])
    return merged


def _write_rows(ws, merged_rows):
    """Clear data rows and rewrite from merged list with sequential Sr. No."""
    # Clear existing data rows (keep header)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for idx, (_, row_data) in enumerate(merged_rows, 1):
        r = idx + 1  # row 2 onwards
        ws.cell(row=r, column=EST_SR_NO).value = idx
        date_cell = ws.cell(row=r, column=EST_DATE)
        date_cell.value = row_data['date']
        date_cell.number_format = 'DD/MM/YYYY'
        ws.cell(row=r, column=EST_NAME).value = row_data['name']
        ws.cell(row=r, column=EST_SIZE).value = row_data['size']
        ws.cell(row=r, column=EST_PROGRAM).value = row_data['program']
        ws.cell(row=r, column=EST_PLY).value = row_data['ply']
        ws.cell(row=r, column=EST_RATE).value = row_data['rate']
        ws.cell(row=r, column=EST_NUM_BOXES).value = row_data['num_boxes']
        ws.cell(row=r, column=EST_TOTAL).value = row_data['total']


def _add_grand_total(ws):
    """Add a Grand Total row below the last data row with a SUM formula."""
    last_data_row = ws.max_row
    if last_data_row < 2:
        return
    total_row = last_data_row + 1
    col_letter = get_column_letter(EST_TOTAL)
    ws.cell(row=total_row, column=EST_NUM_BOXES).value = "Grand Total"
    ws.cell(row=total_row, column=EST_TOTAL).value = (
        f"=SUM({col_letter}2:{col_letter}{last_data_row})"
    )


def transfer(raw_work_path, estimates_path):
    """
    Transfer processed rows from Raw_Work into Estimates.

    Returns a summary dict with counts and per-row details.
    """
    rw_wb = openpyxl.load_workbook(raw_work_path)
    rw_ws = rw_wb.active

    est_wb = openpyxl.load_workbook(estimates_path)

    rows_to_process = []
    for row_num in range(2, rw_ws.max_row + 1):
        name = rw_ws.cell(row=row_num, column=RW_NAME).value
        if name is None:
            break
        rate = rw_ws.cell(row=row_num, column=RW_RATE).value
        total = rw_ws.cell(row=row_num, column=RW_TOTAL).value
        if rate is None or total is None:
            continue
        rows_to_process.append(row_num)

    if not rows_to_process:
        return {'transferred': 0, 'details': [], 'message': 'No rows to transfer (run process first).'}

    # Collect new rows grouped by customer name
    new_rows_by_customer = {}
    details = []

    for row_num in rows_to_process:
        name       = str(rw_ws.cell(row=row_num, column=RW_NAME).value).strip()
        date_val   = rw_ws.cell(row=row_num, column=RW_DATE).value
        length     = float(rw_ws.cell(row=row_num, column=RW_LENGTH).value)
        width      = float(rw_ws.cell(row=row_num, column=RW_WIDTH).value)
        bottom_w   = int(rw_ws.cell(row=row_num, column=RW_BOTTOM_W).value)
        bottom_q   = str(rw_ws.cell(row=row_num, column=RW_BOTTOM_Q).value).strip()
        flute_w    = int(rw_ws.cell(row=row_num, column=RW_FLUTE_W).value)
        flute_q    = str(rw_ws.cell(row=row_num, column=RW_FLUTE_Q).value).strip()
        top_w      = int(rw_ws.cell(row=row_num, column=RW_TOP_W).value)
        top_q      = str(rw_ws.cell(row=row_num, column=RW_TOP_Q).value).strip()
        ply        = int(rw_ws.cell(row=row_num, column=RW_PLY).value)
        order_type = rw_ws.cell(row=row_num, column=RW_ORDER_TYPE).value
        ups        = float(rw_ws.cell(row=row_num, column=RW_UPS).value)
        punching   = rw_ws.cell(row=row_num, column=RW_PUNCHING).value
        pins       = rw_ws.cell(row=row_num, column=RW_PINS).value or 0
        item_name  = rw_ws.cell(row=row_num, column=RW_ITEM_NAME).value
        rate       = rw_ws.cell(row=row_num, column=RW_RATE).value
        num_boxes  = int(rw_ws.cell(row=row_num, column=RW_NUM_BOXES).value)
        total      = rw_ws.cell(row=row_num, column=RW_TOTAL).value

        date_val = _normalize_date(date_val)
        program = _build_program(bottom_w, bottom_q,
                                 flute_w, flute_q, top_w, top_q, ups, punching, pins,
                                 order_type=order_type)
        size = _build_size(length, width)

        row_data = {
            'date': date_val,
            'name': str(item_name).strip() if item_name else '',
            'size': size,
            'program': program,
            'ply': ply,
            'rate': rate,
            'num_boxes': num_boxes,
            'total': total,
        }

        new_rows_by_customer.setdefault(name, []).append((date_val, row_data))

        details.append({
            'row': row_num,
            'name': name,
            'program': program,
            'rate': rate,
            'total': total,
        })

    # For each customer: sort new rows, merge with existing, rewrite sheet
    for customer, new_rows in new_rows_by_customer.items():
        est_ws, _ = _get_or_create_sheet(est_wb, customer)
        existing = _read_existing_rows(est_ws)
        new_rows.sort(key=lambda x: x[0])
        merged = _merge_sorted(existing, new_rows)
        _write_rows(est_ws, merged)
        _add_grand_total(est_ws)

    # Clean up empty default sheet
    if 'Sheet1' in est_wb.sheetnames and len(est_wb.sheetnames) > 1:
        s1 = est_wb['Sheet1']
        if s1.max_row <= 1 and s1.cell(row=1, column=1).value is None:
            est_wb.remove(s1)

    # Remove transferred rows from Raw_Work (reverse order)
    for row_num in reversed(rows_to_process):
        rw_ws.delete_rows(row_num)

    est_wb.save(estimates_path)
    rw_wb.save(raw_work_path)

    return {
        'transferred': len(rows_to_process),
        'details': details,
        'message': f"Transferred {len(rows_to_process)} rows.",
    }
